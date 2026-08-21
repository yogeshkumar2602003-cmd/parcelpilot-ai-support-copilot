"""Deterministic ingestion of ParcelPilot_Assessment_Data.xlsx.

Loads the README snapshot timestamp and the accounts/orders/tickets sheets
verbatim. Nothing here branches on a specific record ID -- rows are loaded
generically by column name.
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import openpyxl

from app.config import DATASET_TZ


@dataclass
class WorkbookData:
    snapshot_at: datetime
    snapshot_raw: str
    currency: str
    accounts: list[dict[str, Any]]
    orders: list[dict[str, Any]]
    tickets: list[dict[str, Any]]


def _sheet_rows(ws) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    out = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        out.append({header[i]: row[i] for i in range(len(header))})
    return out


def _parse_snapshot(raw: str) -> datetime:
    """Parse strings like '2026-08-16 11:00 Asia/Kolkata'. Falls back to
    treating a bare timestamp as DATASET_TZ if no zone is present."""
    m = re.match(r"(\d{4}-\d{2}-\d{2}[ T]\d{2}:\d{2}(?::\d{2})?)\s*([A-Za-z_/]+)?", raw.strip())
    if not m:
        raise ValueError(f"Unrecognized snapshot format in workbook README: {raw!r}")
    ts_part, tz_part = m.group(1), m.group(2)
    ts_part = ts_part.replace("T", " ")
    fmt = "%Y-%m-%d %H:%M:%S" if ts_part.count(":") == 2 else "%Y-%m-%d %H:%M"
    naive = datetime.strptime(ts_part, fmt)
    tz_name = tz_part or DATASET_TZ
    return naive.replace(tzinfo=ZoneInfo(tz_name))


def parse_workbook(path: Path) -> WorkbookData:
    wb = openpyxl.load_workbook(str(path), data_only=True)

    readme_rows = list(wb["README"].iter_rows(values_only=True))
    readme = {str(r[0]).strip(): r[1] for r in readme_rows if r and r[0] is not None}

    snapshot_raw = str(readme.get("Dataset snapshot", "")).strip()
    if not snapshot_raw:
        raise ValueError("Workbook README sheet is missing the 'Dataset snapshot' row")
    snapshot_at = _parse_snapshot(snapshot_raw)
    currency = str(readme.get("Currency", "INR")).strip()

    accounts = _sheet_rows(wb["accounts"])
    orders = _sheet_rows(wb["orders"])
    tickets = _sheet_rows(wb["tickets"])

    return WorkbookData(
        snapshot_at=snapshot_at,
        snapshot_raw=snapshot_raw,
        currency=currency,
        accounts=accounts,
        orders=orders,
        tickets=tickets,
    )
